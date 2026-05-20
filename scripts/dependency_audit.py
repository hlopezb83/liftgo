#!/usr/bin/env python3
"""
LiftGo §20 dependency audit.

Genera dos artefactos a partir de un mapa manual de clasificación
(criterios §20.2-§20.4 de architecture.md):
  - docs/dependency-audit.md (versionado)
  - /mnt/documents/liftgo-dependency-audit.xlsx (descargable)

Los conteos de LOC y consumidores se calculan en vivo sobre el repo.
"""
from __future__ import annotations

import json
import subprocess
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"


# ---------------------------------------------------------------------------
# Clasificación manual (criterios §20)
# ---------------------------------------------------------------------------

@dataclass
class Helper:
    path: str                # relativo a repo
    kind: str                # hook | util | forms | rpc
    canonical: str           # dependencia canónica equivalente, o "—"
    veredict: str            # KEEP / MIGRAR / RETIRADO
    priority: str            # alta / media / baja / —
    action: str

HELPERS: list[Helper] = [
    Helper("src/lib/exportCsv.ts",      "util",  "papaparse",                        "KEEP (ya usa canónica)", "—",     "Ya delega en papaparse; mantener."),
    Helper("src/lib/formatCurrency.ts", "util",  "Intl.NumberFormat (built-in)",     "KEEP (glue <30 LOC)",    "—",     "Wrapper de 27 LOC sobre Intl con locale es-MX."),
    Helper("src/lib/utils.ts",          "util",  "clsx + tailwind-merge + date-fns", "KEEP (ya usa canónicas)","—",     "cn, formatMtyDate, nowMty; delega en deps canónicas."),
    Helper("src/lib/lineItems.ts",      "util",  "—",                                "KEEP (sin equivalente)", "—",     "Narrowing JSONB específico Supabase; no hay equivalente."),
    Helper("src/lib/rpc.ts",            "util",  "@supabase/supabase-js",            "KEEP (glue <30 LOC)",    "—",     "Wrapper tipado de 23 LOC sobre supabase.rpc."),
    Helper("src/lib/telemetry.ts",      "util",  "Sentry (futuro)",                  "KEEP (capa de cambio)",  "baja",  "Abstracción para conectar Sentry sin tocar callers."),
    Helper("src/lib/forms/coerce.ts",   "forms", "—",                                "KEEP (glue <30 LOC)",    "—",     "Coerciones triviales (13 LOC) para form prefill."),
    Helper("src/hooks/use-mobile.tsx",  "hook",  "matchMedia (built-in)",            "KEEP (glue <30 LOC)",    "—",     "Hook breakpoint mobile, 19 LOC."),
    Helper("src/hooks/useDebouncedValue.ts", "hook", "use-debounce",                 "KEEP (glue <30 LOC)",    "baja",  "14 LOC, 1 consumidor. Migrable a use-debounce si crece."),
    Helper("src/hooks/useDialogState.ts","hook", "—",                                "KEEP (sin equivalente)", "—",     "Patrón propio para Sheet/Dialog state."),
    Helper("src/hooks/useFormState.ts", "hook",  "react-hook-form",                  "MIGRAR",                 "media", "9 LOC, 7 dialogs. Duplica RHF (ya canónico); migrar incrementalmente."),
    Helper("src/hooks/useListFilters.ts","hook", "react-router + @tanstack/react-table","KEEP (ya usa canónicas)","—",  "Compone useSearchParams + filtro client; no duplica."),
    Helper("src/hooks/useListPage.ts",  "hook",  "@tanstack/react-table",            "KEEP (ya usa canónica)", "—",     "Headless 100% TanStack Table; documentado en §20.4."),
    Helper("src/hooks/useDocuments.ts", "hook",  "@tanstack/react-query",            "KEEP (ya usa canónica)", "—",     "Domain hook sobre supabase + react-query."),
    Helper("src/hooks/useBreadcrumbEntityLabel.ts","hook","@tanstack/react-query",   "KEEP (ya usa canónica)", "—",     "Domain hook."),
    Helper("src/hooks/useVisibleNavGroups.ts","hook","—",                            "KEEP (sin equivalente)", "—",     "Lógica de visibilidad por rol."),
]

# ---------------------------------------------------------------------------
# Stack canónico §20.4
# ---------------------------------------------------------------------------

CANONICAL = {
    "@tanstack/react-query":  "Estado servidor",
    "@tanstack/react-table":  "Tablas",
    "@tanstack/react-virtual":"Tablas (virtual)",
    "@hookform/resolvers":    "Formularios",
    "react-hook-form":        "Formularios",
    "zod":                    "Validación",
    "date-fns":               "Fechas / zonas horarias",
    "date-fns-tz":            "Fechas / zonas horarias",
    "@react-pdf/renderer":    "PDF",
    "papaparse":              "CSV",
    "sonner":                 "Toasts",
    "react-dropzone":         "Drag & drop archivos",
    "clsx":                   "Class merging",
    "tailwind-merge":         "Class merging",
    "lucide-react":           "Iconos",
    "currency.js":            "Cálculos financieros",
    "tailwindcss-animate":    "Animaciones",
    "class-variance-authority":"UI primitives (shadcn)",
    "@supabase/supabase-js":  "Backend (Cloud)",
    "react-router-dom":       "Router",
    "react": "UI runtime", "react-dom": "UI runtime",
    "next-themes": "Tema",
    "cmdk": "UI primitives (shadcn)",
    "react-day-picker": "Fechas (calendar UI)",
    "recharts": "Charts",
    "vaul": "UI primitives (shadcn)",
    "file-saver": "Descargas blob (lazy import)",
    "html2canvas": "Captura screenshot DOM (lazy, feedback)",
}

# Radix primitives en bloque
RADIX_PREFIX = "@radix-ui/"

# Dependencias que NO están en §20.4 (a evaluar)
NON_CANONICAL_NOTES = {
    "dompurify": "Sanitiza HTML del manual (help system). Usado puntualmente.",
    "@hello-pangea/dnd": "Drag & drop para Kanban de feedback/maintenance. Canónica de facto.",
}


def run(cmd: list[str]) -> str:
    return subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, check=False).stdout


def loc(path: str) -> int:
    p = ROOT / path
    return len(p.read_text().splitlines()) if p.exists() else 0


def consumers(path: str) -> int:
    """Cuenta archivos que importan este módulo (excluyendo el archivo mismo)."""
    # Convierte src/foo/bar.ts -> @/foo/bar
    rel = path.removeprefix("src/").rsplit(".", 1)[0]
    pattern = f'from "@/{rel}"'
    out = run(["rg", "-l", "--fixed-strings", pattern, "src"])
    files = [f for f in out.strip().splitlines() if f and f != path]
    return len(files)


def load_deps() -> dict[str, str]:
    pkg = json.loads((ROOT / "package.json").read_text())
    return pkg.get("dependencies", {})


def classify_dep(name: str) -> tuple[str, str]:
    if name in CANONICAL:
        return "Canónica activa", CANONICAL[name]
    if name.startswith(RADIX_PREFIX):
        return "Canónica activa", "UI primitives (shadcn/Radix)"
    return "No canónica (evaluar)", NON_CANONICAL_NOTES.get(name, "Revisar uso real y justificación §20.3.")


# ---------------------------------------------------------------------------
# Build dataset
# ---------------------------------------------------------------------------

helper_rows = []
for h in HELPERS:
    helper_rows.append({
        "Archivo":             h.path,
        "Tipo":                h.kind,
        "LOC":                 loc(h.path),
        "Consumidores":        consumers(h.path),
        "Dep canónica equiv.": h.canonical,
        "Veredicto":           h.veredict,
        "Prioridad":           h.priority,
        "Acción":              h.action,
    })

deps = load_deps()
dep_rows = []
for name, version in sorted(deps.items()):
    cat, note = classify_dep(name)
    static = run(["rg", "-l", "--fixed-strings", f'from "{name}"', "src"])
    dynamic = run(["rg", "-l", "--fixed-strings", f'import("{name}")', "src"])
    files = set(f for f in static.strip().splitlines() if f) | set(
        f for f in dynamic.strip().splitlines() if f
    )
    dep_rows.append({
        "Paquete":     name,
        "Versión":     version,
        "Categoría":   cat,
        "Mapeo §20.4": note,
        "Consumidores": len(files),
    })

# KPIs
total_helpers     = len(helper_rows)
veredict_counts   = {}
for r in helper_rows:
    veredict_counts[r["Veredicto"]] = veredict_counts.get(r["Veredicto"], 0) + 1
loc_total         = sum(r["LOC"] for r in helper_rows)
loc_migrar        = sum(r["LOC"] for r in helper_rows if r["Veredicto"].startswith("MIGRAR"))
deps_canonical    = sum(1 for r in dep_rows if r["Categoría"] == "Canónica activa")
deps_non_canonical= sum(1 for r in dep_rows if r["Categoría"] == "No canónica (evaluar)")

summary_rows = [
    ("Helpers auditados",                total_helpers),
    ("Helpers KEEP (glue/sin equiv.)",   sum(v for k, v in veredict_counts.items() if k.startswith("KEEP"))),
    ("Helpers a MIGRAR",                 veredict_counts.get("MIGRAR", 0)),
    ("LOC propio total",                 loc_total),
    ("LOC potencialmente migrable",      loc_migrar),
    ("Dependencias canónicas activas",   deps_canonical),
    ("Dependencias no canónicas",        deps_non_canonical),
]


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------

def md_table(headers: list[str], rows: list[list]) -> str:
    out = ["| " + " | ".join(headers) + " |",
           "| " + " | ".join(["---"] * len(headers)) + " |"]
    for r in rows:
        out.append("| " + " | ".join(str(c) for c in r) + " |")
    return "\n".join(out)


md = []
md.append("# Auditoría §20 — Dependencias vs helpers internos\n")
md.append(f"> Generado por `scripts/dependency_audit.py`. Doctrina: `architecture.md` §20.\n")
md.append("## Resumen ejecutivo\n")
md.append(f"- Helpers internos auditados: **{total_helpers}** ({loc_total} LOC propios).")
md.append(f"- KEEP: **{sum(v for k,v in veredict_counts.items() if k.startswith('KEEP'))}** · MIGRAR: **{veredict_counts.get('MIGRAR',0)}** · RETIRADO: **{veredict_counts.get('RETIRADO',0)}**.")
md.append(f"- LOC potencialmente migrable a una dep canónica: **{loc_migrar}**.")
md.append(f"- Dependencias canónicas activas (§20.4): **{deps_canonical}** · no canónicas a evaluar: **{deps_non_canonical}**.")
md.append("- Aplicaciones previas de §20 ya entregadas: `6.6.0-alpha.1` (PDF → @react-pdf/renderer), `6.6.0-alpha.3` (toasts → sonner), `6.6.0-alpha.4` (PR template).\n")

md.append("## Tabla 1 — Helpers internos\n")
md.append(md_table(
    list(helper_rows[0].keys()),
    [list(r.values()) for r in helper_rows],
))

md.append("\n## Tabla 2 — Dependencias (`package.json`)\n")
md.append(md_table(
    list(dep_rows[0].keys()),
    [list(r.values()) for r in dep_rows],
))

md.append("\n## Oportunidades priorizadas\n")
migrar = [r for r in helper_rows if r["Veredicto"].startswith("MIGRAR")]
non_canon = [r for r in dep_rows if r["Categoría"] == "No canónica (evaluar)"]
md.append("### Helpers a migrar")
if migrar:
    for r in migrar:
        md.append(f"- **{r['Archivo']}** → `{r['Dep canónica equiv.']}` (prioridad {r['Prioridad']}, {r['Consumidores']} consumidores). {r['Acción']}")
else:
    md.append("- Ninguno pendiente.")

md.append("\n### Dependencias no canónicas a evaluar")
if non_canon:
    for r in non_canon:
        md.append(f"- **{r['Paquete']}@{r['Versión']}** ({r['Consumidores']} consumidores) — {r['Mapeo §20.4']}")
else:
    md.append("- Ninguna.")

md.append("\n## Historial de aplicaciones de §20\n")
md.append("| Versión | Cambio | Resultado |")
md.append("| --- | --- | --- |")
md.append("| 6.6.0-alpha.1 | jsPDF → @react-pdf/renderer | PDFs declarativos en todos los documentos |")
md.append("| 6.6.0-alpha.2 | Doctrina §20 documentada | architecture.md §20.1–§20.7 |")
md.append("| 6.6.0-alpha.3 | Toast legacy → sonner | -186 LOC, -1 dep (@radix-ui/react-toast) |")
md.append("| 6.6.0-alpha.4 | PR template con checklist §20 | .github/pull_request_template.md |")
md.append("")

(ROOT / "docs" / "dependency-audit.md").write_text("\n".join(md))


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

wb = Workbook()
HEADER_FILL = PatternFill("solid", start_color="FFFF00")
BOLD = Font(name="Arial", bold=True)
BODY = Font(name="Arial")


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, h in enumerate(headers, start=1):
        max_len = max([len(str(h))] + [len(str(r[col_idx - 1])) for r in rows if col_idx - 1 < len(r)])
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(14, max_len + 2), 60)
    ws.freeze_panes = "A2"


ws1 = wb.active
ws1.title = "Helpers"
write_sheet(ws1, list(helper_rows[0].keys()), [list(r.values()) for r in helper_rows])

ws2 = wb.create_sheet("Dependencias")
write_sheet(ws2, list(dep_rows[0].keys()), [list(r.values()) for r in dep_rows])

ws3 = wb.create_sheet("Resumen")
write_sheet(ws3, ["Métrica", "Valor"], [list(r) for r in summary_rows])

out_xlsx = Path("/mnt/documents/liftgo-dependency-audit.xlsx")
out_xlsx.parent.mkdir(parents=True, exist_ok=True)
wb.save(out_xlsx)

print(f"OK · docs/dependency-audit.md ({total_helpers} helpers, {len(dep_rows)} deps)")
print(f"OK · {out_xlsx}")
